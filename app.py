import re
import os
import secrets
from io import BytesIO
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from datetime import date, datetime, time, timedelta
from functools import wraps

import pytz
from flask import (
    Flask,
    abort,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import or_
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from werkzeug.security import check_password_hash, generate_password_hash
from sqlalchemy import case, func, inspect, text
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm import joinedload

app = Flask(__name__)
uri = os.environ.get("DATABASE_URL", "").strip()
if uri.startswith("postgres://"):
    uri = uri.replace("postgres://", "postgresql://", 1)
if not uri:
    os.makedirs(os.path.join(app.root_path, "instance"), exist_ok=True)
    uri = "sqlite:///" + os.path.join(app.root_path, "instance", "altai.db").replace("\\", "/")
    app.logger.warning("DATABASE_URL is not set; using local SQLite at instance/altai.db")
app.config["SQLALCHEMY_DATABASE_URI"] = uri
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
_secret = os.getenv("SECRET_KEY")
if not _secret:
    _secret = secrets.token_hex(32)
    app.logger.warning("SECRET_KEY is not set; sessions reset on every restart")
app.config["SECRET_KEY"] = _secret
db = SQLAlchemy(app)
TZ = pytz.timezone(os.getenv("TZ", "Asia/Bishkek"))
UTC = pytz.UTC


def _db_stores_utc_naive():
    uri = app.config.get("SQLALCHEMY_DATABASE_URI", "")
    return uri.startswith("postgresql")


def _now_for_db():
    if _db_stores_utc_naive():
        return datetime.now(UTC).replace(tzinfo=None)
    return datetime.now(TZ).replace(tzinfo=None)

CLIENT_AUTH_MAX_ATTEMPTS = 5
CLIENT_AUTH_WINDOW_SECONDS = 300
_client_auth_attempts = {}

CSRF_EXEMPT_ENDPOINTS = frozenset({"login", "client_auth", "client_portal_login", "idle_logout", "static"})
SESSION_IDLE_MINUTES = int(os.getenv("SESSION_IDLE_MINUTES", "10"))
SESSION_IDLE_SECONDS = SESSION_IDLE_MINUTES * 60
SESSION_IDLE_EXEMPT_ENDPOINTS = frozenset(
    {"static", "login", "client_auth", "client_portal_login", "idle_logout"}
)

def _to_local_datetime(dt):
    if not dt:
        return None
    if dt.tzinfo is not None:
        return dt.astimezone(TZ)
    if _db_stores_utc_naive():
        return UTC.localize(dt).astimezone(TZ)
    return TZ.localize(dt)


def _local_to_db_datetime(local_dt):
    if local_dt.tzinfo is None:
        local_dt = TZ.localize(local_dt)
    else:
        local_dt = local_dt.astimezone(TZ)
    if _db_stores_utc_naive():
        return local_dt.astimezone(UTC).replace(tzinfo=None)
    return local_dt.replace(tzinfo=None)


def fmt_dt_local(dt):
    local_dt = _to_local_datetime(dt)
    if not local_dt:
        return "-"
    return local_dt.strftime("%d.%m.%Y %H:%M")


app.jinja_env.filters["dt_kz"] = fmt_dt_local


def fmt_dt_local_input(dt):
    local_dt = _to_local_datetime(dt)
    if not local_dt:
        return ""
    return local_dt.strftime("%Y-%m-%dT%H:%M")


app.jinja_env.filters["dt_local_input"] = fmt_dt_local_input


def fmt_dt_local_date(dt):
    local_dt = _to_local_datetime(dt)
    if not local_dt:
        return "-"
    return local_dt.strftime("%d.%m.%Y")


app.jinja_env.filters["dt_kz_date"] = fmt_dt_local_date


def fmt_phone(phone):
    """Format a 10-digit phone string as XXXX-XXX-XXX."""
    digits = re.sub(r"\D", "", str(phone or ""))
    if len(digits) == 10:
        return f"{digits[:4]}-{digits[4:7]}-{digits[7:]}"
    return phone or ""


app.jinja_env.filters["fmt_phone"] = fmt_phone


def fmt_money(value):
    if value is None or value == "":
        return "—"
    try:
        return f"{float(value):.2f} сом"
    except (TypeError, ValueError):
        return str(value)


app.jinja_env.filters["fmt_money"] = fmt_money


def ensure_csrf_token():
    if "csrf_token" not in session:
        session["csrf_token"] = secrets.token_urlsafe(32)
    return session["csrf_token"]


def validate_csrf_token(token=None):
    token = token or request.form.get("csrf_token") or request.headers.get("X-CSRF-Token", "")
    expected = session.get("csrf_token", "")
    return bool(token and expected and secrets.compare_digest(token, expected))


def _is_client_auth_rate_limited():
    ip = request.remote_addr or "unknown"
    now = datetime.now(TZ).timestamp()
    attempts, window_start = _client_auth_attempts.get(ip, (0, now))
    if now - window_start > CLIENT_AUTH_WINDOW_SECONDS:
        return False
    return attempts >= CLIENT_AUTH_MAX_ATTEMPTS


def _record_failed_client_auth():
    ip = request.remote_addr or "unknown"
    now = datetime.now(TZ).timestamp()
    attempts, window_start = _client_auth_attempts.get(ip, (0, now))
    if now - window_start > CLIENT_AUTH_WINDOW_SECONDS:
        attempts, window_start = 0, now
    _client_auth_attempts[ip] = (attempts + 1, window_start)


def _reset_client_auth_rate_limit():
    ip = request.remote_addr or "unknown"
    _client_auth_attempts.pop(ip, None)


def _session_is_authenticated():
    return bool(session.get("user_id")) or session.get("role") == "client"


def _update_session_activity():
    session["last_activity"] = datetime.now(TZ).timestamp()


def _session_idle_expired():
    last_activity = session.get("last_activity")
    if last_activity is None:
        return False
    return datetime.now(TZ).timestamp() - float(last_activity) > SESSION_IDLE_SECONDS


def _release_client_unique_fields(client):
    stamp = int(datetime.now(TZ).timestamp())
    client.fio = f"{client.fio[:72]}__del{client.id}_{stamp}"[:100]
    client.phone = f"9{client.id:09d}"[-10:]
    client.inn = f"9{client.id:013d}"[-14:]
    client.token = None


def _release_car_unique_fields(car):
    car.number = f"DEL{car.id:07d}"[:20]


def _adjust_daily_stock(liters_delta, stock_date=None):
    _ensure_daily_stock_tables()
    stock_date = stock_date or datetime.now(TZ).date()
    daily_stock = _get_or_create_daily_stock(stock_date)
    daily_stock.current_stock = _to_decimal_2(
        _to_decimal_2(daily_stock.current_stock or 0) + _to_decimal_2(liters_delta)
    )


def _parse_local_datetime(value):
    if not value:
        return None
    naive = datetime.strptime(value, "%Y-%m-%dT%H:%M")
    return _local_to_db_datetime(TZ.localize(naive))


@app.context_processor
def inject_csrf():
    if session.get("user_id") or session.get("role") == "client":
        ensure_csrf_token()
    return {
        "csrf_token": session.get("csrf_token", ""),
        "session_idle_minutes": SESSION_IDLE_MINUTES,
    }


@app.before_request
def session_idle_timeout():
    if request.endpoint in SESSION_IDLE_EXEMPT_ENDPOINTS:
        return
    if not _session_is_authenticated():
        return
    if _session_idle_expired():
        was_client = session.get("role") == "client"
        session.clear()
        flash(
            f"Вы вышли из системы: {SESSION_IDLE_MINUTES} минут бездействия.",
            "warning",
        )
        return redirect(url_for("client_auth") if was_client else url_for("login"))
    _update_session_activity()


@app.before_request
def ensure_active_client_session():
    if request.endpoint in SESSION_IDLE_EXEMPT_ENDPOINTS:
        return
    if session.get("role") != "client":
        return
    client_id = session.get("client_id")
    if not client_id:
        return
    client = Client.query.get(client_id)
    if not client or client.is_deleted:
        session.clear()
        flash("Доступ в кабинет закрыт.", "warning")
        return redirect(url_for("client_auth"))


@app.before_request
def csrf_protect():
    if request.method not in ("POST", "PUT", "DELETE", "PATCH"):
        return
    if request.endpoint in CSRF_EXEMPT_ENDPOINTS:
        return
    if not session.get("user_id") and session.get("role") != "client":
        return
    ensure_csrf_token()
    if validate_csrf_token():
        return
    ensure_csrf_token()
    if request.path.startswith("/api/") or request.is_json:
        return jsonify({"ok": False, "error": "Обновите страницу и повторите действие."}), 403
    flash("Не удалось отправить форму. Обновите страницу и попробуйте снова.", "warning")
    return redirect(request.referrer or url_for("index"))


class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(50), unique=True, nullable=False)
    password = db.Column(db.String(255), nullable=False)
    role = db.Column(db.String(20), nullable=False, default="operator")
    created_at = db.Column(db.DateTime, default=_now_for_db)


class Client(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    fio = db.Column(db.String(100), unique=True, nullable=False)
    phone = db.Column(db.String(10), nullable=False)
    inn = db.Column(db.String(14), nullable=False)
    is_deleted = db.Column(db.Boolean, default=False, server_default=text("false"), nullable=False)
    deleted_at = db.Column(db.DateTime, nullable=True)
    created_at = db.Column(db.DateTime, default=_now_for_db)
    cars = db.relationship("Car", backref="client", lazy=True)
    token = db.Column(db.String(64), unique=True)

    def generate_token(self):
        self.token = secrets.token_urlsafe(32)


class Car(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    client_id = db.Column(db.Integer, db.ForeignKey("client.id"), nullable=False)
    number = db.Column(db.String(20), unique=True, nullable=False)
    brand = db.Column(db.String(100), nullable=False)
    color = db.Column(db.String(50), nullable=False)
    note = db.Column(db.Text, nullable=True)
    stock = db.Column(db.Numeric(10, 2), default=0)
    is_deleted = db.Column(db.Boolean, default=False, server_default=text("false"), nullable=False)
    deleted_at = db.Column(db.DateTime, nullable=True)
    created_at = db.Column(db.DateTime, default=_now_for_db)
    sales = db.relationship("Sale", backref="car", lazy=True)
    receipts = db.relationship("Receipt", backref="car", lazy=True)


PAYMENT_METHODS = ["долг", "наличка", "безнал", "доллар"]


class Sale(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    car_id = db.Column(db.Integer, db.ForeignKey("car.id"), nullable=False)
    liters = db.Column(db.Numeric(10, 2), nullable=False)
    price_per_liter = db.Column(db.Numeric(10, 2), nullable=False)
    total = db.Column(db.Numeric(10, 2), nullable=False)
    payment_method = db.Column(db.String(20), nullable=False)
    payment_amount = db.Column(db.Numeric(10, 2), nullable=True)
    created_by = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)
    note = db.Column(db.Text, nullable=True)
    created_at = db.Column(db.DateTime, default=_now_for_db)
    payments = db.relationship("Payment", backref="sale", lazy=True, cascade="all, delete-orphan")
    created_by_user = db.relationship("User", foreign_keys=[created_by], lazy="joined")


class Receipt(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    car_id = db.Column(db.Integer, db.ForeignKey("car.id"), nullable=False)
    liters = db.Column(db.Numeric(10, 2), nullable=False)
    amount = db.Column(db.Numeric(10, 2), nullable=False)
    created_at = db.Column(db.DateTime, default=_now_for_db)
    notes = db.Column(db.Text, nullable=True)


class DailyStock(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    stock_date = db.Column(db.Date, nullable=False, unique=True)
    current_stock = db.Column(db.Numeric(10, 2), default=0)
    created_at = db.Column(db.DateTime, default=_now_for_db)


class StockHistory(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    stock_date = db.Column(db.Date, nullable=False)
    added_liters = db.Column(db.Numeric(10, 2), nullable=False)
    created_at = db.Column(db.DateTime, default=_now_for_db)


PAYMENT_TYPES = ["продажа", "долг"]
DEBT_PAYMENT_TYPE = "долг"

REPORT_TYPE_LABELS = {
    "sales": "Журнал продаж",
    "payments": "Журнал оплат",
    "cash": "Касса",
    "turnover": "Оборот",
}


class Payment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    client_id = db.Column(db.Integer, db.ForeignKey("client.id"), nullable=False)
    sale_id = db.Column(db.Integer, db.ForeignKey("sale.id"), nullable=True)
    amount = db.Column(db.Numeric(10, 2), nullable=False)
    payment_type = db.Column(
        db.String(20),
        db.CheckConstraint("payment_type IN ('продажа', 'долг')"),
        nullable=False,
    )
    payment_method = db.Column(db.String(20), nullable=True)
    paid_by = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)
    created_at = db.Column(db.DateTime, default=_now_for_db)
    client = db.relationship("Client", backref=db.backref("payments", lazy=True))
    paid_by_user = db.relationship("User", foreign_keys=[paid_by], lazy="joined")


def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated_function


def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "user_id" not in session or session.get("role") != "admin":
            return redirect(url_for("index"))
        return f(*args, **kwargs)
    return decorated_function


def staff_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "user_id" not in session or session.get("role") not in ("admin", "operator"):
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated_function


def verify_user_password(user, password):
    if not user.password:
        return False
    return check_password_hash(user.password, password)


def validate_phone(phone):
    digits = re.sub(r"\D", "", phone)
    return digits if len(digits) == 10 else None


def validate_inn(inn):
    digits = re.sub(r"\D", "", inn)
    return digits if len(digits) == 14 else None


def _parse_iso_date(value):
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return None


def _coerce_day(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        try:
            return datetime.strptime(value, "%Y-%m-%d").date()
        except ValueError:
            return None
    return None


def _parse_month_value(month_value):
    if not month_value:
        return None
    try:
        return datetime.strptime(month_value, "%Y-%m").date().replace(day=1)
    except ValueError:
        return None


def _month_bounds(month_value=None):
    month_start = _parse_month_value(month_value)
    if not month_start:
        today = datetime.now(TZ).date()
        month_start = today.replace(day=1)
    next_month_start = (month_start.replace(day=28) + timedelta(days=4)).replace(day=1)
    month_end = next_month_start - timedelta(days=1)
    return month_start, month_end, month_start.strftime("%Y-%m")


def _journal_month_bounds():
    """Текущий и предыдущий календарные месяцы для журналов."""
    today = datetime.now(TZ).date()
    current_start = today.replace(day=1)
    previous_start = (current_start - timedelta(days=1)).replace(day=1)
    next_month_start = (current_start.replace(day=28) + timedelta(days=4)).replace(day=1)
    current_end = next_month_start - timedelta(days=1)
    return previous_start, current_end


def _journal_period_label(start_date, end_date):
    if start_date.replace(day=1) == end_date.replace(day=1):
        return _month_label(start_date)
    return f"{_month_label(start_date)} — {_month_label(end_date)}"


def _month_label(month_start):
    months_ru = [
        "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
        "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь",
    ]
    return f"{months_ru[month_start.month - 1]} {month_start.year}"


def _month_label_filename(month_start):
    return _month_label(month_start).replace(" ", "")


def _format_number(value):
    return round(float(value or 0), 2)


def _to_decimal_2(value):
    if value is None:
        return Decimal("0.00")
    return Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _month_datetime_bounds(start_date, end_date):
    start_local = TZ.localize(datetime.combine(start_date, time.min))
    end_local = TZ.localize(datetime.combine(end_date + timedelta(days=1), time.min))
    if _db_stores_utc_naive():
        return (
            start_local.astimezone(UTC).replace(tzinfo=None),
            end_local.astimezone(UTC).replace(tzinfo=None),
        )
    return start_local, end_local


def _local_date(dt):
    local_dt = _to_local_datetime(dt)
    return local_dt.date() if local_dt else None


def _ensure_daily_stock_tables():
    inspector = inspect(db.engine)
    table_names = set(inspector.get_table_names())
    if "daily_stock" not in table_names:
        DailyStock.__table__.create(bind=db.engine, checkfirst=True)
    if "stock_history" not in table_names:
        StockHistory.__table__.create(bind=db.engine, checkfirst=True)


def _ensure_client_token_column():
    inspector = inspect(db.engine)
    if "client" not in set(inspector.get_table_names()):
        return
    columns = {column["name"] for column in inspector.get_columns("client")}
    if "token" not in columns:
        db.session.execute(text("ALTER TABLE client ADD COLUMN token VARCHAR(64)"))
        db.session.commit()
        return
    token_column = next(
        (column for column in inspector.get_columns("client") if column["name"] == "token"),
        None,
    )
    if not token_column:
        return
    column_type = str(token_column["type"]).lower()
    if "64" in column_type or "text" in column_type:
        return
    if db.engine.dialect.name == "postgresql":
        db.session.execute(text("ALTER TABLE client ALTER COLUMN token TYPE VARCHAR(64)"))
    else:
        db.session.execute(text("ALTER TABLE client ADD COLUMN token__tmp VARCHAR(64)"))
        db.session.execute(text("UPDATE client SET token__tmp = token"))
        db.session.execute(text("ALTER TABLE client DROP COLUMN token"))
        db.session.execute(text("ALTER TABLE client RENAME COLUMN token__tmp TO token"))
    db.session.commit()


def _client_portal_url(token):
    return url_for("client_portal_login", token=token, _external=True)


def _ensure_client_car_soft_delete():
    inspector = inspect(db.engine)
    table_names = set(inspector.get_table_names())

    for table_name in ["client", "car"]:
        if table_name in table_names:
            columns = {column["name"] for column in inspector.get_columns(table_name)}
            if "is_deleted" not in columns:
                if table_name == "client":
                    db.session.execute(text("ALTER TABLE client ADD COLUMN is_deleted BOOLEAN DEFAULT 0"))
                else:
                    db.session.execute(text("ALTER TABLE car ADD COLUMN is_deleted BOOLEAN DEFAULT 0"))
            if "deleted_at" not in columns:
                if table_name == "client":
                    db.session.execute(text("ALTER TABLE client ADD COLUMN deleted_at DATETIME"))
                else:
                    db.session.execute(text("ALTER TABLE car ADD COLUMN deleted_at DATETIME"))
    db.session.commit()


def _ensure_sale_payment_user_columns():
    inspector = inspect(db.engine)
    table_names = set(inspector.get_table_names())
    if "sale" in table_names:
        sale_columns = {column["name"] for column in inspector.get_columns("sale")}
        if "created_by" not in sale_columns:
            db.session.execute(text("ALTER TABLE sale ADD COLUMN created_by INTEGER"))
            db.session.commit()
    if "payment" in table_names:
        payment_columns = {column["name"] for column in inspector.get_columns("payment")}
        if "paid_by" not in payment_columns:
            db.session.execute(text("ALTER TABLE payment ADD COLUMN paid_by INTEGER"))
            db.session.commit()


def _ensure_liters_numeric_columns():
    inspector = inspect(db.engine)
    table_names = set(inspector.get_table_names())

    def _needs_numeric_migration(table_name):
        column = next(
            (col for col in inspector.get_columns(table_name) if col["name"] == "liters"),
            None,
        )
        if not column:
            return False
        column_type = str(column["type"]).lower()
        return "float" in column_type or "real" in column_type or "double" in column_type

    def _sqlite_copy_sale_to_numeric():
        sale_columns = {column["name"] for column in inspector.get_columns("sale")}
        if "created_by" not in sale_columns:
            db.session.execute(text("ALTER TABLE sale ADD COLUMN created_by INTEGER"))
        if "note" not in sale_columns:
            db.session.execute(text("ALTER TABLE sale ADD COLUMN note TEXT"))
        if "created_at" not in sale_columns:
            db.session.execute(text("ALTER TABLE sale ADD COLUMN created_at DATETIME"))

        db.session.execute(text("DROP TABLE IF EXISTS sale__tmp"))
        db.session.execute(
            text(
                """
                CREATE TABLE sale__tmp (
                    id INTEGER NOT NULL PRIMARY KEY,
                    car_id INTEGER NOT NULL,
                    liters NUMERIC(10, 2) NOT NULL,
                    price_per_liter NUMERIC(10, 2) NOT NULL,
                    total NUMERIC(10, 2) NOT NULL,
                    payment_method VARCHAR(20) NOT NULL,
                    payment_amount NUMERIC(10, 2),
                    created_by INTEGER,
                    note TEXT,
                    created_at DATETIME,
                    FOREIGN KEY(car_id) REFERENCES car (id),
                    FOREIGN KEY(created_by) REFERENCES user (id)
                )
                """
            )
        )
        db.session.execute(
            text(
                """
                INSERT INTO sale__tmp
                    (id, car_id, liters, price_per_liter, total, payment_method, payment_amount, created_by, note, created_at)
                SELECT
                    id,
                    car_id,
                    CAST(liters AS NUMERIC),
                    price_per_liter,
                    total,
                    payment_method,
                    payment_amount,
                    created_by,
                    note,
                    created_at
                FROM sale
                """
            )
        )
        db.session.execute(text("DROP TABLE sale"))
        db.session.execute(text("ALTER TABLE sale__tmp RENAME TO sale"))

    def _sqlite_copy_receipt_to_numeric():
        receipt_columns = {column["name"] for column in inspector.get_columns("receipt")}
        if "created_at" not in receipt_columns:
            db.session.execute(text("ALTER TABLE receipt ADD COLUMN created_at DATETIME"))
        if "notes" not in receipt_columns:
            db.session.execute(text("ALTER TABLE receipt ADD COLUMN notes TEXT"))

        db.session.execute(text("DROP TABLE IF EXISTS receipt__tmp"))
        db.session.execute(
            text(
                """
                CREATE TABLE receipt__tmp (
                    id INTEGER NOT NULL PRIMARY KEY,
                    car_id INTEGER NOT NULL,
                    liters NUMERIC(10, 2) NOT NULL,
                    amount NUMERIC(10, 2) NOT NULL,
                    created_at DATETIME,
                    notes TEXT,
                    FOREIGN KEY(car_id) REFERENCES car (id)
                )
                """
            )
        )
        db.session.execute(
            text(
                """
                INSERT INTO receipt__tmp (id, car_id, liters, amount, created_at, notes)
                SELECT
                    id,
                    car_id,
                    CAST(liters AS NUMERIC),
                    amount,
                    created_at,
                    notes
                FROM receipt
                """
            )
        )
        db.session.execute(text("DROP TABLE receipt"))
        db.session.execute(text("ALTER TABLE receipt__tmp RENAME TO receipt"))

    sale_needs_migration = "sale" in table_names and _needs_numeric_migration("sale")
    receipt_needs_migration = "receipt" in table_names and _needs_numeric_migration("receipt")
    if not sale_needs_migration and not receipt_needs_migration:
        return

    if db.engine.dialect.name == "sqlite":
        db.session.execute(text("PRAGMA foreign_keys=OFF"))
        migration_step = "initialization"
        try:
            if sale_needs_migration:
                migration_step = "sale.liters"
                _sqlite_copy_sale_to_numeric()
                db.session.commit()
            if receipt_needs_migration:
                migration_step = "receipt.liters"
                _sqlite_copy_receipt_to_numeric()
                db.session.commit()
        except SQLAlchemyError as exc:
            db.session.rollback()
            raise RuntimeError(f"Failed to migrate {migration_step} column to NUMERIC(10, 2).") from exc
        finally:
            db.session.execute(text("PRAGMA foreign_keys=ON"))
            db.session.commit()
    else:
        if sale_needs_migration:
            db.session.execute(
                text("ALTER TABLE sale ALTER COLUMN liters TYPE NUMERIC(10, 2)")
            )
        if receipt_needs_migration:
            db.session.execute(
                text("ALTER TABLE receipt ALTER COLUMN liters TYPE NUMERIC(10, 2)")
            )
        db.session.commit()


def _get_or_create_daily_stock(stock_date):
    daily_stock = DailyStock.query.filter_by(stock_date=stock_date).first()
    if daily_stock:
        return daily_stock

    previous_day_stock = (
        DailyStock.query
        .filter(DailyStock.stock_date < stock_date)
        .order_by(DailyStock.stock_date.desc())
        .first()
    )
    base_stock = _to_decimal_2(previous_day_stock.current_stock) if previous_day_stock else Decimal("0.00")
    daily_stock = DailyStock(stock_date=stock_date, current_stock=base_stock)
    db.session.add(daily_stock)
    db.session.flush()
    return daily_stock


def _bootstrap_database():
    db.create_all()
    _ensure_daily_stock_tables()
    _ensure_client_token_column()
    _ensure_client_car_soft_delete()
    _ensure_sale_payment_user_columns()
    _ensure_liters_numeric_columns()

    for legacy_user in User.query.all():
        if not (legacy_user.password or "").startswith(("pbkdf2:", "scrypt:")):
            legacy_user.password = generate_password_hash(legacy_user.password or "")

    if not User.query.filter_by(username="admin").first():
        admin_password = os.getenv("DEFAULT_ADMIN_PASSWORD")
        if admin_password:
            db.session.add(
                User(username="admin", password=generate_password_hash(admin_password), role="admin")
            )
            app.logger.info("Created default admin user")
        else:
            app.logger.warning(
                "User 'admin' not found and DEFAULT_ADMIN_PASSWORD is not set; "
                "create an admin manually or set DEFAULT_ADMIN_PASSWORD"
            )

    if not User.query.filter_by(username="operator").first():
        operator_password = os.getenv("DEFAULT_OPERATOR_PASSWORD")
        if operator_password:
            db.session.add(
                User(
                    username="operator",
                    password=generate_password_hash(operator_password),
                    role="operator",
                )
            )
            app.logger.info("Created default operator user")
        else:
            app.logger.warning(
                "User 'operator' not found and DEFAULT_OPERATOR_PASSWORD is not set"
            )

    db.session.commit()


with app.app_context():
    _bootstrap_database()


@app.route("/login", methods=["GET", "POST"])
def login():
    if "user_id" in session:
        return redirect(url_for("index"))

    if request.method == "POST":
        ensure_csrf_token()
        if not validate_csrf_token():
            return render_template("login.html", error="Сессия истекла. Повторите вход.")

        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()

        user = User.query.filter_by(username=username).first()
        if user and verify_user_password(user, password):
            session.clear()
            session["user_id"] = user.id
            session["username"] = user.username
            session["role"] = user.role
            ensure_csrf_token()
            _update_session_activity()
            return redirect(url_for("index"))

        error = "Неверное имя пользователя или пароль"
        return render_template("login.html", error=error)

    ensure_csrf_token()
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/idle-logout")
def idle_logout():
    was_client = session.get("role") == "client"
    session.clear()
    flash(
        f"Вы вышли из системы: {SESSION_IDLE_MINUTES} минут бездействия.",
        "warning",
    )
    return redirect(url_for("client_auth") if was_client else url_for("login"))


@app.route("/")
@login_required
def index():
    if session.get("role") == "admin":
        return redirect(url_for("clients"))
    return redirect(url_for("sales"))


@app.route("/clients")
@login_required
def clients():
    q = request.args.get("q", "").strip()
    query = Client.query.filter_by(is_deleted=False)
    if q:
        query = query.filter(
            or_(
                Client.fio.ilike(f"%{q}%"),
                Client.phone.ilike(f"%{q}%")
            )
        )
    all_clients = query.order_by(Client.id.desc()).all()
    return render_template("clients.html", clients=all_clients, q=q)

@app.route("/delete-sale/<int:id>", methods=["POST"])
@admin_required
def delete_sale(id):
    sale = Sale.query.filter_by(id=id).first_or_404()
    sale_date = _local_date(sale.created_at) or datetime.now(TZ).date()
    liters = sale.liters or 0
    db.session.delete(sale)
    _adjust_daily_stock(liters, sale_date)
    db.session.commit()
    return redirect(url_for("sales_journal"))

@app.route("/admin/users")
@admin_required
def admin_users():
    users = User.query.order_by(User.id.asc()).all()
    return render_template("admin/users.html", users=users)


@app.route("/admin/users/add", methods=["GET", "POST"])
@admin_required
def admin_add_user():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        role = request.form.get("role", "operator")
        if role not in {"operator", "admin"}:
            role = "operator"

        if not username or not password:
            flash("Заполните все поля!", "danger")
            return redirect(url_for("admin_add_user"))

        if User.query.filter_by(username=username).first():
            flash("Пользователь уже существует!", "danger")
            return redirect(url_for("admin_add_user"))

        user = User(username=username, password=generate_password_hash(password), role=role)
        db.session.add(user)
        db.session.commit()
        flash(f"Пользователь {username} создан!", "success")
        return redirect(url_for("admin_users"))

    return render_template("admin/add_user.html")


@app.route("/admin/users/<int:user_id>/edit", methods=["GET", "POST"])
@admin_required
def admin_edit_user(user_id):
    user = User.query.get_or_404(user_id)

    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        role = request.form.get("role", "operator")
        if role not in {"operator", "admin"}:
            role = "operator"

        if not username:
            flash("Заполните имя пользователя!", "danger")
            return redirect(url_for("admin_edit_user", user_id=user_id))

        if username != user.username and User.query.filter_by(username=username).first():
            flash("Имя пользователя уже используется!", "danger")
            return redirect(url_for("admin_edit_user", user_id=user_id))

        user.username = username
        if password:
            user.password = generate_password_hash(password)
        user.role = role
        db.session.commit()
        flash("Пользователь обновлен!", "success")
        return redirect(url_for("admin_users"))

    return render_template("admin/edit_user.html", user=user)


@app.route("/admin/users/<int:user_id>/delete", methods=["POST"])
@admin_required
def admin_delete_user(user_id):
    user = User.query.get_or_404(user_id)
    if user.id == session.get("user_id"):
        flash("Нельзя удалить свой аккаунт!", "danger")
        return redirect(url_for("admin_users"))

    username = user.username
    db.session.delete(user)
    db.session.commit()
    flash(f"Пользователь {username} удален!", "success")
    return redirect(url_for("admin_users"))


def _sale_remaining_debt(sale):
    total = sale.total if sale.total is not None else (sale.price_per_liter or 0) * (sale.liters or 0)
    paid = sale.payment_amount or 0
    return max(Decimal("0"), _to_decimal_2(total) - _to_decimal_2(paid))


def _recalculate_sale_payment_amount(sale):
    total_paid = sum(
        (_to_decimal_2(payment.amount) for payment in sale.payments),
        Decimal("0"),
    )
    sale.payment_amount = total_paid if total_paid > 0 else None
    return sale.payment_amount


def _sync_sale_payments_after_edit(sale):
    debt_paid = sum(
        (
            _to_decimal_2(payment.amount)
            for payment in sale.payments
            if payment.payment_type == "долг"
        ),
        Decimal("0"),
    )
    for payment in list(sale.payments):
        if payment.payment_type == "продажа":
            db.session.delete(payment)

    total_paid = _to_decimal_2(sale.payment_amount or 0)
    sale_part = max(Decimal("0"), total_paid - debt_paid)
    if sale.payment_method != "долг" and sale_part > 0:
        db.session.add(
            Payment(
                client_id=sale.car.client_id,
                sale_id=sale.id,
                amount=sale_part,
                payment_type="продажа",
                payment_method=sale.payment_method,
                paid_by=session.get("user_id"),
            )
        )
    _recalculate_sale_payment_amount(sale)


@app.route("/client-dashboard")
def client_dashboard():
    if session.get("role") != "client":
        return redirect(url_for("client_auth"))
    client_id = session.get("client_id")
    client = Client.query.filter_by(id=client_id, is_deleted=False).first_or_404()
    cars = Car.query.filter_by(client_id=client_id, is_deleted=False).order_by(Car.number.asc()).all()
    sales = (
        Sale.query.options(joinedload(Sale.car))
        .join(Car)
        .filter(Car.client_id == client_id, Car.is_deleted.is_(False))
        .order_by(Sale.created_at.desc())
        .all()
    )
    debt_sales = [sale for sale in sales if _sale_remaining_debt(sale) > 0]
    payments = (
        Payment.query.filter_by(client_id=client_id)
        .options(joinedload(Payment.sale))
        .order_by(Payment.created_at.desc())
        .all()
    )
    total_debt = sum((_sale_remaining_debt(sale) for sale in sales), Decimal("0"))
    return render_template(
        "client_dashboard.html",
        client=client,
        sales=sales,
        debt_sales=debt_sales,
        payments=payments,
        cars=cars,
        total_debt=float(total_debt),
    )


@app.route("/client-logout", methods=["POST"])
def client_logout():
    session.clear()
    return redirect(url_for("client_auth"))


@app.route("/add-client", methods=["GET", "POST"])
@admin_required
def add_client():
    errors = {}
    form = {}
    if request.method == "POST":
        form["fio"] = request.form.get("fio", "").strip()
        form["phone"] = request.form.get("phone", "").strip()
        form["inn"] = request.form.get("inn", "").strip()
        form["car_number"] = request.form.get("car_number", "").strip()
        form["car_brand"] = request.form.get("car_brand", "").strip()
        form["car_color"] = request.form.get("car_color", "").strip()
        form["car_note"] = request.form.get("car_note", "").strip()

        if not form["fio"]:
            errors["fio"] = "ФИО обязательно."
        elif Client.query.filter_by(fio=form["fio"], is_deleted=False).first():
            errors["fio"] = "Клиент с таким ФИО уже существует."

        phone_digits = validate_phone(form["phone"])
        if phone_digits is None:
            errors["phone"] = "Телефон должен содержать ровно 10 цифр."
        elif Client.query.filter_by(phone=phone_digits, is_deleted=False).first():
            errors["phone"] = "Телефон уже используется."

        inn_digits = validate_inn(form["inn"])
        if inn_digits is None:
            errors["inn"] = "ИНН должен содержать ровно 14 цифр."
        elif Client.query.filter_by(inn=inn_digits, is_deleted=False).first():
            errors["inn"] = "ИНН уже используется."

        if not form["car_number"]:
            errors["car_number"] = "Номер машины обязателен."
        elif Car.query.filter_by(number=form["car_number"], is_deleted=False).first():
            errors["car_number"] = "Машина с таким номером уже существует."

        if not form["car_brand"]:
            errors["car_brand"] = "Марка машины обязательна."

        if not form["car_color"]:
            errors["car_color"] = "Цвет машины обязателен."

        if not errors:
            client = Client(fio=form["fio"], phone=phone_digits, inn=inn_digits)
            db.session.add(client)
            db.session.flush()
            car = Car(
                client_id=client.id,
                number=form["car_number"],
                brand=form["car_brand"],
                color=form["car_color"],
                note=form["car_note"] or None,
            )
            db.session.add(car)
            db.session.commit()
            return redirect(url_for("client_detail", id=client.id))

    return render_template("add_client.html", errors=errors, form=form)


@app.route("/client/<int:id>")
@login_required
def client_detail(id):
    client = Client.query.filter_by(id=id, is_deleted=False).first_or_404()
    cars = Car.query.filter_by(client_id=id, is_deleted=False).order_by(Car.id.desc()).all()
    return render_template("client_detail.html", client=client, cars=cars)


@app.route("/edit-client/<int:id>", methods=["GET", "POST"])
@admin_required
def edit_client(id):
    client = Client.query.filter_by(id=id, is_deleted=False).first_or_404()
    errors = {}
    form = {
        "fio": client.fio,
        "phone": client.phone,
        "inn": client.inn,
    }
    if request.method == "POST":
        form["fio"] = request.form.get("fio", "").strip()
        form["phone"] = request.form.get("phone", "").strip()
        form["inn"] = request.form.get("inn", "").strip()

        if not form["fio"]:
            errors["fio"] = "ФИО обязательно."
        else:
            existing = Client.query.filter_by(fio=form["fio"], is_deleted=False).first()
            if existing and existing.id != id:
                errors["fio"] = "Клиент с таким ФИО уже существует."

        phone_digits = validate_phone(form["phone"])
        if phone_digits is None:
            errors["phone"] = "Телефон должен содержать ровно 10 цифр."
        else:
            existing_phone = Client.query.filter_by(phone=phone_digits, is_deleted=False).first()
            if existing_phone and existing_phone.id != id:
                errors["phone"] = "Телефон уже используется."

        inn_digits = validate_inn(form["inn"])
        if inn_digits is None:
            errors["inn"] = "ИНН должен содержать ровно 14 цифр."
        else:
            existing_inn = Client.query.filter_by(inn=inn_digits, is_deleted=False).first()
            if existing_inn and existing_inn.id != id:
                errors["inn"] = "ИНН уже используется."

        if not errors:
            client.fio = form["fio"]
            client.phone = phone_digits
            client.inn = inn_digits
            db.session.commit()
            return redirect(url_for("client_detail", id=id))

    return render_template("edit_client.html", client=client, errors=errors, form=form)


def _login_client_by_token(token):
    token = (token or "").strip()
    if not token:
        return None
    return Client.query.filter_by(token=token, is_deleted=False).first()


def _start_client_session(client):
    session.clear()
    session["client_id"] = client.id
    session["role"] = "client"
    ensure_csrf_token()
    _update_session_activity()


@app.route("/client-cabinet/<token>")
def client_portal_login(token):
    client = _login_client_by_token(token)
    if not client:
        return render_template(
            "client_auth.html",
            error="Ссылка недействительна или устарела. Запросите новую у менеджера.",
        ), 404
    _start_client_session(client)
    return redirect(url_for("client_dashboard"))


@app.route("/generate-token", methods=["GET", "POST"])
@login_required
@staff_required
def generate_token():
    client_id = request.args.get("client_id", type=int)
    if not client_id:
        flash("Не указан клиент.", "danger")
        return redirect(url_for("clients"))
    client = Client.query.filter_by(id=client_id, is_deleted=False).first_or_404()
    portal_url = _client_portal_url(client.token) if client.token else None
    if request.method == "POST":
        client.generate_token()
        db.session.commit()
        portal_url = _client_portal_url(client.token)
        flash("Ссылка для клиента создана.", "success")
    return render_template(
        "generate_token.html",
        client=client,
        portal_url=portal_url,
    )


@app.route("/client-auth", methods=["GET", "POST"])
def client_auth():
    if request.method == "POST":
        if _is_client_auth_rate_limited():
            return render_template(
                "client_auth.html",
                error="Слишком много попыток. Подождите несколько минут.",
            )
        token = request.form.get("token", "").strip()
        client = _login_client_by_token(token)
        if client:
            _reset_client_auth_rate_limit()
            _start_client_session(client)
            return redirect(url_for("client_dashboard"))
        _record_failed_client_auth()
        return render_template("client_auth.html", error="Ссылка или код доступа неверны")
    return render_template("client_auth.html")


@app.route("/delete-client/<int:id>", methods=["POST"])
@admin_required
def delete_client(id):
    client = Client.query.filter_by(id=id, is_deleted=False).first_or_404()
    deleted_at = _now_for_db()
    client.is_deleted = True
    client.deleted_at = deleted_at
    _release_client_unique_fields(client)
    for car in Car.query.filter_by(client_id=client.id, is_deleted=False).all():
        car.is_deleted = True
        car.deleted_at = deleted_at
        _release_car_unique_fields(car)
    db.session.commit()
    return redirect(url_for("clients"))


@app.route("/add-car", methods=["GET", "POST"])
@admin_required
def add_car():
    all_clients = Client.query.filter_by(is_deleted=False).order_by(Client.fio).all()
    errors = {}
    form = {}
    if request.method == "POST":
        form["client_id"] = request.form.get("client_id", "").strip()
        form["number"] = request.form.get("number", "").strip()
        form["brand"] = request.form.get("brand", "").strip()
        form["color"] = request.form.get("color", "").strip()
        form["note"] = request.form.get("note", "").strip()

        if not form["client_id"]:
            errors["client_id"] = "Выберите клиента."

        if not form["number"]:
            errors["number"] = "Номер машины обязателен."
        elif Car.query.filter_by(number=form["number"], is_deleted=False).first():
            errors["number"] = "Машина с таким номером уже существует."

        if not form["brand"]:
            errors["brand"] = "Марка машины обязательна."

        if not form["color"]:
            errors["color"] = "Цвет машины обязателен."

        if not errors:
            try:
                selected_client_id = int(form["client_id"])
            except (TypeError, ValueError):
                selected_client_id = None
            client = None
            if selected_client_id:
                client = Client.query.filter_by(id=selected_client_id, is_deleted=False).first()
            if not client:
                errors["client_id"] = "Выберите корректного клиента."
            else:
                car = Car(
                    client_id=client.id,
                    number=form["number"],
                    brand=form["brand"],
                    color=form["color"],
                    note=form["note"] or None,
                )
                db.session.add(car)
                db.session.commit()
                return redirect(url_for("client_detail", id=car.client_id))

    return render_template("add_car.html", clients=all_clients, errors=errors, form=form)


@app.route("/edit-car/<int:id>", methods=["GET", "POST"])
@admin_required
def edit_car(id):
    car = Car.query.filter_by(id=id, is_deleted=False).first_or_404()
    errors = {}
    form = {
        "number": car.number,
        "brand": car.brand,
        "color": car.color,
        "note": car.note or "",
    }
    if request.method == "POST":
        form["number"] = request.form.get("number", "").strip()
        form["brand"] = request.form.get("brand", "").strip()
        form["color"] = request.form.get("color", "").strip()
        form["note"] = request.form.get("note", "").strip()

        if not form["number"]:
            errors["number"] = "Номер машины обязателен."
        else:
            existing = Car.query.filter_by(number=form["number"], is_deleted=False).first()
            if existing and existing.id != id:
                errors["number"] = "Машина с таким номером уже существует."

        if not form["brand"]:
            errors["brand"] = "Марка машины обязательна."

        if not form["color"]:
            errors["color"] = "Цвет машины обязателен."

        if not errors:
            car.number = form["number"]
            car.brand = form["brand"]
            car.color = form["color"]
            car.note = form["note"] or None
            db.session.commit()
            return redirect(url_for("client_detail", id=car.client_id))

    return render_template("edit_car.html", car=car, errors=errors, form=form)


@app.route("/cars")
@login_required
def cars():
    q = request.args.get("q", "").strip()
    query = (
        Car.query.filter_by(is_deleted=False)
        .options(joinedload(Car.client))
        .join(Car.client)
        .filter(Client.is_deleted.is_(False))
    )
    if q:
        query = query.filter(
            or_(
                Client.fio.ilike(f"%{q}%"),
                Car.number.ilike(f"%{q}%")
            )
        )
    all_cars = query.order_by(Car.id.desc()).all()
    return render_template("cars.html", cars=all_cars, q=q)


@app.route("/delete-car/<int:id>", methods=["POST"])
@admin_required
def delete_car(id):
    car = Car.query.filter_by(id=id, is_deleted=False).first_or_404()
    client_id = car.client_id
    next_page = request.form.get("next", "")
    car.is_deleted = True
    car.deleted_at = _now_for_db()
    _release_car_unique_fields(car)
    db.session.commit()
    if next_page == "cars":
        return redirect(url_for("cars"))
    return redirect(url_for("client_detail", id=client_id))


@app.route("/api/check-fio", methods=["POST"])
@admin_required
def check_fio():
    payload = request.get_json(silent=True) or {}
    fio = payload.get("fio", "").strip()
    if not fio:
        return jsonify({"exists": False})

    existing = Client.query.filter_by(fio=fio, is_deleted=False).first()
    return jsonify({"exists": bool(existing)})


@app.route("/api/check-phone", methods=["POST"])
@admin_required
def check_phone():
    payload = request.get_json(silent=True) or {}
    phone = payload.get("phone", "").strip()
    client_id = payload.get("client_id")
    digits = re.sub(r"\D", "", phone)
    is_valid = len(digits) == 10

    exists = False
    if is_valid:
        query = Client.query.filter_by(phone=digits, is_deleted=False)
        if client_id is not None and client_id != "":
            try:
                query = query.filter(Client.id != int(client_id))
            except (TypeError, ValueError):
                pass
        exists = bool(query.first())

    return jsonify({
        "valid": is_valid,
        "exists": exists,
        "formatted": fmt_phone(digits) if is_valid else "",
    })


@app.route("/api/check-inn", methods=["POST"])
@admin_required
def check_inn():
    payload = request.get_json(silent=True) or {}
    inn = payload.get("inn", "").strip()
    client_id = payload.get("client_id")
    digits = re.sub(r"\D", "", inn)
    is_valid = len(digits) == 14

    exists = False
    if is_valid:
        query = Client.query.filter_by(inn=digits, is_deleted=False)
        if client_id is not None and client_id != "":
            try:
                query = query.filter(Client.id != int(client_id))
            except (TypeError, ValueError):
                pass
        exists = bool(query.first())

    return jsonify({"valid": is_valid, "exists": exists})


@app.route("/api/check-car-number", methods=["POST"])
@admin_required
def check_car_number():
    payload = request.get_json(silent=True) or {}
    car_number = payload.get("car_number", "").strip()
    car_id = payload.get("car_id")
    if not car_number:
        return jsonify({"exists": False})

    query = Car.query.filter_by(number=car_number, is_deleted=False)
    if car_id is not None and car_id != "":
        try:
            query = query.filter(Car.id != int(car_id))
        except (TypeError, ValueError):
            pass
    existing = query.first()
    return jsonify({"exists": bool(existing)})


@app.route("/api/set-daily-stock", methods=["POST"])
@admin_required
def set_daily_stock():
    _ensure_daily_stock_tables()

    payload = request.get_json(silent=True) or {}
    liters_raw = payload.get("liters")

    try:
        liters = Decimal(str(liters_raw))
        if liters <= 0:
            raise ValueError
    except (TypeError, ValueError, InvalidOperation):
        return jsonify({"ok": False, "error": "Количество литров должно быть больше 0."}), 400

    today = datetime.now(TZ).date()
    daily_stock = _get_or_create_daily_stock(today)
    liters = _to_decimal_2(liters)
    daily_stock.current_stock = _to_decimal_2(daily_stock.current_stock or 0) + liters
    db.session.add(StockHistory(stock_date=today, added_liters=liters))
    db.session.commit()

    app.logger.info(
        "Daily stock increased: stock_date=%s, liters=%s, user_id=%s",
        today.isoformat(),
        str(liters),
        session.get("user_id"),
    )

    return jsonify({
        "ok": True,
        "stock_date": today.isoformat(),
        "current_stock": str(_to_decimal_2(daily_stock.current_stock)),
        "added_liters": str(liters),
    })


@app.route("/api/car-search")
@login_required
def api_car_search():
    q = request.args.get("q", "").strip()
    if not q:
        return jsonify([])
    cars = (
        Car.query.options(joinedload(Car.client))
        .join(Car.client)
        .filter(Car.is_deleted.is_(False), Client.is_deleted.is_(False), Car.number.ilike(f"%{q}%"))
        .limit(10)
        .all()
    )
    return jsonify([
        {"id": c.id, "number": c.number, "fio": c.client.fio}
        for c in cars
    ])


@app.route("/receipts", methods=["GET", "POST"])
@login_required
def receipts():
    start_date = _parse_iso_date(request.args.get("start_date"))
    end_date = _parse_iso_date(request.args.get("end_date"))
    errors = {}
    form = {}

    if request.method == "POST":
        form["car_id"] = request.form.get("car_id", "").strip()
        form["liters"] = request.form.get("liters", "").strip()
        form["amount"] = request.form.get("amount", "").strip()
        form["notes"] = request.form.get("notes", "").strip()

        car = None
        if not form["car_id"]:
            errors["car_id"] = "Выберите машину."
        else:
            try:
                car = Car.query.get(int(form["car_id"]))
                if not car or car.is_deleted or (car.client and car.client.is_deleted):
                    errors["car_id"] = "Машина не найдена."
            except (TypeError, ValueError):
                errors["car_id"] = "Выберите корректную машину."

        try:
            liters = float(form["liters"])
            if liters <= 0:
                raise ValueError
        except (ValueError, TypeError):
            errors["liters"] = "Введите корректное количество литров."
            liters = None

        try:
            amount = Decimal(form["amount"])
            if amount <= 0:
                raise ValueError
            amount = _to_decimal_2(amount)
        except (ValueError, TypeError, InvalidOperation):
            errors["amount"] = "Введите корректную сумму."
            amount = None

        if not errors:
            db.session.add(
                Receipt(
                    car_id=car.id,
                    liters=liters,
                    amount=amount,
                    notes=form["notes"] or None,
                )
            )
            _adjust_daily_stock(liters)
            db.session.commit()
            return redirect(url_for("receipts"))

    cars = (
        Car.query.options(joinedload(Car.client))
        .join(Car.client)
        .filter(Car.is_deleted.is_(False), Client.is_deleted.is_(False))
        .order_by(Car.number.asc())
        .all()
    )
    query = Receipt.query.options(joinedload(Receipt.car).joinedload(Car.client))
    if start_date and end_date:
        min_dt, max_dt = _month_datetime_bounds(start_date, end_date)
        query = query.filter(Receipt.created_at >= min_dt, Receipt.created_at < max_dt)
    elif start_date:
        min_dt, max_dt = _month_datetime_bounds(start_date, start_date)
        query = query.filter(Receipt.created_at >= min_dt, Receipt.created_at < max_dt)
    elif end_date:
        min_dt, max_dt = _month_datetime_bounds(end_date, end_date)
        query = query.filter(Receipt.created_at >= min_dt, Receipt.created_at < max_dt)
    all_receipts = query.order_by(Receipt.created_at.desc()).all()

    return render_template(
        "receipts.html",
        receipts=all_receipts,
        cars=cars,
        errors=errors,
        form=form,
        start_date=start_date.isoformat() if start_date else "",
        end_date=end_date.isoformat() if end_date else "",
    )


@app.route("/receipts/<int:receipt_id>", methods=["POST", "DELETE"])
@admin_required
def delete_receipt(receipt_id):
    receipt = Receipt.query.get_or_404(receipt_id)
    receipt_date = _local_date(receipt.created_at) or datetime.now(TZ).date()
    liters = receipt.liters or 0
    db.session.delete(receipt)
    _adjust_daily_stock(-Decimal(str(liters)), receipt_date)
    db.session.commit()
    return redirect(url_for("receipts"))


@app.route("/sales", methods=["GET", "POST"])
@login_required
def sales():
    errors = {}
    form = {}
    client_info = None
    if request.method == "POST":
        form["car_number"] = request.form.get("car_number", "").strip()
        form["liters"] = request.form.get("liters", "").strip()
        form["price_per_liter"] = request.form.get("price_per_liter", "").strip()
        form["payment_method"] = request.form.get("payment_method", "").strip()
        form["payment_amount"] = request.form.get("payment_amount", "").strip()
        form["note"] = request.form.get("note", "").strip()

        car = None
        found_car = (
            Car.query
            .options(joinedload(Car.client))
            .filter(Car.number == form["car_number"])
            .first()
        ) if form["car_number"] else None
        if not form["car_number"]:
            errors["car_number"] = "Введите номер машины."
        elif not found_car:
            errors["car_number"] = "Машина с таким номером не найдена."
        elif found_car.is_deleted:
            errors["car_number"] = "Эта машина удалена и больше не доступна."
        elif not found_car.client:
            errors["car_number"] = "Эта машина не связана с клиентом."
        elif found_car.client.is_deleted:
            errors["car_number"] = "Этот клиент удален и больше не доступен."
        else:
            car = found_car

        try:
            liters = float(form["liters"])
            if liters <= 0:
                raise ValueError
        except (ValueError, TypeError):
            errors["liters"] = "Введите корректное количество литров."
            liters = None

        try:
            price = Decimal(form["price_per_liter"])
            if price <= 0:
                raise ValueError
            price = _to_decimal_2(price)
        except (ValueError, TypeError, InvalidOperation):
            errors["price_per_liter"] = "Введите корректную цену за литр."
            price = None

        if form["payment_method"] not in PAYMENT_METHODS:
            errors["payment_method"] = "Выберите способ оплаты."

        payment_amount = None
        if form["payment_method"] != "долг":
            if not form["payment_amount"]:
                errors["payment_amount"] = "Введите сумму оплаты."
            else:
                try:
                    payment_amount = Decimal(form["payment_amount"])
                    if payment_amount < 0:
                        raise ValueError
                    payment_amount = _to_decimal_2(payment_amount)
                except (ValueError, TypeError, InvalidOperation):
                    errors["payment_amount"] = "Введите корректную сумму оплаты."

        if not errors and form["payment_method"] != "долг" and payment_amount is not None:
            expected_total = _to_decimal_2(Decimal(str(liters)) * price)
            if payment_amount > expected_total:
                errors["payment_amount"] = (
                    f"Сумма оплаты не может превышать итог ({expected_total})."
                )

        if not errors:
            total = _to_decimal_2(Decimal(str(liters)) * price)
            sale = Sale(
                car_id=car.id,
                liters=liters,
                price_per_liter=price,
                total=total,
                payment_method=form["payment_method"],
                payment_amount=payment_amount,
                created_by=session.get("user_id"),
                note=form["note"] or None,
            )
            db.session.add(sale)
            db.session.flush()
            if form["payment_method"] != "долг" and payment_amount:
                payment = Payment(
                    client_id=car.client_id,
                    sale_id=sale.id,
                    amount=payment_amount,
                    payment_type="продажа",
                    payment_method=form["payment_method"],
                    paid_by=session.get("user_id"),
                )
                db.session.add(payment)
            _ensure_daily_stock_tables()
            sale_date = datetime.now(TZ).date()
            daily_stock = _get_or_create_daily_stock(sale_date)
            daily_stock.current_stock = _to_decimal_2(
                _to_decimal_2(daily_stock.current_stock or 0) - _to_decimal_2(liters)
            )
            db.session.commit()
            return redirect(url_for("sales_journal"))

        if car:
            client_info = {"fio": car.client.fio}

    return render_template(
        "sales.html",
        errors=errors,
        form=form,
        client_info=client_info,
        payment_methods=PAYMENT_METHODS,
    )


def _build_sales_report_payload(start_date, end_date):
    start_dt, end_dt = _month_datetime_bounds(start_date, end_date)
    sales = (
        Sale.query
        .options(
            joinedload(Sale.car).joinedload(Car.client),
            joinedload(Sale.created_by_user),
        )
        .join(Sale.car)
        .join(Car.client)
        .filter(
            Sale.created_at >= start_dt,
            Sale.created_at < end_dt,
            Car.is_deleted.is_(False),
            Client.is_deleted.is_(False),
        )
        .order_by(Sale.created_at.desc())
        .all()
    )
    rows = []
    for sale in sales:
        payment_amount = _format_number(sale.payment_amount) if sale.payment_amount is not None else None
        total = _format_number(sale.total)
        remaining = _format_number(total - (payment_amount or 0))
        rows.append(
            {
                "date": fmt_dt_local(sale.created_at),
                "client": sale.car.client.fio,
                "car_number": sale.car.number,
                "liters": _format_number(sale.liters),
                "price_per_liter": _format_number(sale.price_per_liter),
                "total": total,
                "payment_amount": payment_amount if payment_amount is not None else "—",
                "payment_method": sale.payment_method or "—",
                "operator": sale.created_by_user.username if sale.created_by_user else "—",
                "remaining": remaining,
                "note": sale.note or "—",
            }
        )
    return {
        "columns": [
            {"key": "date", "label": "Дата"},
            {"key": "client", "label": "Клиент"},
            {"key": "car_number", "label": "Номер машины"},
            {"key": "liters", "label": "Литры"},
            {"key": "price_per_liter", "label": "Цена / л"},
            {"key": "total", "label": "Сумма, сом"},
            {"key": "payment_amount", "label": "Оплата, сом"},
            {"key": "payment_method", "label": "Способ оплаты"},
            {"key": "operator", "label": "Оператор"},
            {"key": "remaining", "label": "Остаток, сом"},
            {"key": "note", "label": "Примечание"},
        ],
        "rows": rows,
    }


def _build_payments_report_payload(start_date, end_date):
    start_dt, end_dt = _month_datetime_bounds(start_date, end_date)
    payments_data = (
        Payment.query
        .options(
            joinedload(Payment.sale).joinedload(Sale.car),
            joinedload(Payment.client),
            joinedload(Payment.paid_by_user),
        )
        .join(Payment.client)
        .filter(
            Payment.created_at >= start_dt,
            Payment.created_at < end_dt,
            Client.is_deleted.is_(False),
        )
        .order_by(Payment.created_at.desc())
        .all()
    )
    rows = []
    for payment in payments_data:
        sale = payment.sale
        remaining = _format_number(_sale_remaining_debt(sale)) if sale else "0"
        payment_method = payment.payment_method or (sale.payment_method if sale else None)
        rows.append(
            {
                "sale_date": fmt_dt_local(sale.created_at) if sale else "—",
                "client": payment.client.fio,
                "car_number": sale.car.number if sale and sale.car else "—",
                "liters": _format_number(sale.liters) if sale else "—",
                "price_per_liter": _format_number(sale.price_per_liter) if sale else "—",
                "total": _format_number(sale.total) if sale else "—",
                "payment_amount": _format_number(payment.amount),
                "payment_method": payment_method or "—",
                "operator": payment.paid_by_user.username if payment.paid_by_user else "—",
                "remaining": remaining,
                "payment_date": fmt_dt_local(payment.created_at),
            }
        )
    return {
        "columns": [
            {"key": "sale_date", "label": "Дата продажи"},
            {"key": "client", "label": "Клиент"},
            {"key": "car_number", "label": "Номер машины"},
            {"key": "liters", "label": "Литры"},
            {"key": "price_per_liter", "label": "Цена / л"},
            {"key": "total", "label": "Сумма"},
            {"key": "payment_amount", "label": "Сумма оплаты"},
            {"key": "payment_method", "label": "Способ оплаты"},
            {"key": "operator", "label": "Оператор"},
            {"key": "remaining", "label": "Остаток суммы"},
            {"key": "payment_date", "label": "Дата и время оплаты"},
        ],
        "rows": rows,
    }


def _build_cash_rows(start_date, end_date):
    start_dt, end_dt = _month_datetime_bounds(start_date, end_date)
    all_payments = (
        Payment.query
        .filter(Payment.created_at >= start_dt, Payment.created_at < end_dt)
        .order_by(Payment.created_at.desc())
        .all()
    )
    daily = {}
    for payment in all_payments:
        date_key = _local_date(payment.created_at)
        if not date_key:
            continue
        if date_key not in daily:
            daily[date_key] = {
                "date": date_key.strftime("%d.%m.%Y"),
                "sale_наличка": 0.0,
                "sale_безнал": 0.0,
                "sale_доллар": 0.0,
                "debt_наличка": 0.0,
                "debt_безнал": 0.0,
                "debt_доллар": 0.0,
            }
        method = payment.payment_method or ""
        if payment.payment_type == "продажа" and method in ("наличка", "безнал", "доллар"):
            daily[date_key][f"sale_{method}"] += float(payment.amount or 0)
        elif payment.payment_type == "долг" and method in ("наличка", "безнал", "доллар"):
            daily[date_key][f"debt_{method}"] += float(payment.amount or 0)

    rows = []
    for date_key, data in daily.items():
        row = dict(data)
        row["total_наличка"] = _format_number(row["sale_наличка"] + row["debt_наличка"])
        row["total_безнал"] = _format_number(row["sale_безнал"] + row["debt_безнал"])
        row["total_доллар"] = _format_number(row["sale_доллар"] + row["debt_доллар"])
        row["sale_наличка"] = _format_number(row["sale_наличка"])
        row["sale_безнал"] = _format_number(row["sale_безнал"])
        row["sale_доллар"] = _format_number(row["sale_доллар"])
        row["debt_наличка"] = _format_number(row["debt_наличка"])
        row["debt_безнал"] = _format_number(row["debt_безнал"])
        row["debt_доллар"] = _format_number(row["debt_доллар"])
        row["_sort_date"] = date_key
        rows.append(row)
    rows.sort(key=lambda row: row["_sort_date"], reverse=True)
    for row in rows:
        row.pop("_sort_date", None)
    return rows


def _build_cash_report_payload(start_date, end_date):
    return {
        "columns": [
            {"key": "date", "label": "Дата"},
            {"key": "sale_наличка", "label": "Наличка сегодня"},
            {"key": "sale_безнал", "label": "Безнал сегодня"},
            {"key": "sale_доллар", "label": "Доллар сегодня"},
            {"key": "debt_наличка", "label": "Наличка от долгов"},
            {"key": "debt_безнал", "label": "Безнал от долгов"},
            {"key": "debt_доллар", "label": "Доллар от долгов"},
            {"key": "total_наличка", "label": "Наличка всего"},
            {"key": "total_безнал", "label": "Безнал всего"},
            {"key": "total_доллар", "label": "Доллар всего"},
        ],
        "rows": _build_cash_rows(start_date, end_date),
    }


def _build_turnover_rows(start_date, end_date):
    _ensure_daily_stock_tables()
    min_dt, max_dt = _month_datetime_bounds(start_date, end_date)

    rows_data = []
    totals = {"liters": 0.0, "amount": 0.0, "payments": 0.0, "debts": 0.0, "average_price": 0.0}
    error_message = None

    try:
        sales_by_day = {}
        for sale in Sale.query.filter(
            Sale.created_at >= min_dt,
            Sale.created_at < max_dt,
        ).all():
            row_date = _local_date(sale.created_at)
            if not row_date:
                continue
            if row_date not in sales_by_day:
                sales_by_day[row_date] = {
                    "liters": 0.0,
                    "amount": 0.0,
                    "sale_ids": [],
                }
            sales_by_day[row_date]["liters"] += float(sale.liters or 0)
            sales_by_day[row_date]["amount"] += float(sale.total or 0)
            sales_by_day[row_date]["sale_ids"].append(sale.id)

        for row_date, daily_sales in sales_by_day.items():
            liters = daily_sales["liters"]
            amount = daily_sales["amount"]
            sale_ids = daily_sales["sale_ids"]
            if sale_ids:
                payments = float(
                    db.session.query(func.coalesce(func.sum(Payment.amount), 0))
                    .filter(Payment.sale_id.in_(sale_ids))
                    .scalar()
                    or 0
                )
            else:
                payments = 0.0
            debts = max(0.0, amount - payments)
            average_price = amount / liters if liters else 0.0
            totals["liters"] += liters
            totals["amount"] += amount
            totals["payments"] += payments
            totals["debts"] += debts
            sales_by_day[row_date] = {
                "liters": liters,
                "amount": amount,
                "payments": payments,
                "debts": debts,
                "average_price": average_price,
            }

        additions_by_day = {
            _coerce_day(row.stock_date): float(row.added_liters or 0)
            for row in db.session.query(
                StockHistory.stock_date.label("stock_date"),
                func.coalesce(func.sum(StockHistory.added_liters), 0).label("added_liters"),
            ).filter(StockHistory.stock_date >= start_date, StockHistory.stock_date <= end_date).group_by(StockHistory.stock_date).all()
            if _coerce_day(row.stock_date)
        }
        stock_by_day = {
            _coerce_day(row.stock_date): float(row.current_stock or 0)
            for row in db.session.query(
                DailyStock.stock_date.label("stock_date"),
                DailyStock.current_stock.label("current_stock"),
            ).filter(DailyStock.stock_date >= start_date, DailyStock.stock_date <= end_date).all()
            if _coerce_day(row.stock_date)
        }

        all_days = set(sales_by_day) | set(additions_by_day) | set(stock_by_day)
        for row_date in sorted(all_days, reverse=True):
            daily_sales = sales_by_day.get(
                row_date,
                {"liters": 0.0, "amount": 0.0, "payments": 0.0, "debts": 0.0, "average_price": 0.0},
            )
            rows_data.append(
                {
                    "date": row_date,
                    "date_label": row_date.strftime("%d.%m.%Y"),
                    "liters": daily_sales["liters"],
                    "amount": daily_sales["amount"],
                    "payments": daily_sales["payments"],
                    "debts": daily_sales["debts"],
                    "average_price": daily_sales["average_price"],
                    "remaining_goods": stock_by_day.get(row_date, 0.0),
                    "added_liters": additions_by_day.get(row_date, 0.0),
                }
            )
        totals["average_price"] = totals["amount"] / totals["liters"] if totals["liters"] else 0.0
    except SQLAlchemyError:
        app.logger.exception("Failed to build turnover report for date range %s - %s", start_date.isoformat(), end_date.isoformat())
        rows_data = []
        error_message = "Не удалось загрузить данные оборота. Проверьте подключение к базе данных."
    return rows_data, totals, error_message


def _build_turnover_report_payload(start_date, end_date):
    rows_data, totals, error_message = _build_turnover_rows(start_date, end_date)
    return {
        "columns": [
            {"key": "date_label", "label": "Дата"},
            {"key": "liters", "label": "Литр"},
            {"key": "amount", "label": "Сумма"},
            {"key": "payments", "label": "Оплаты"},
            {"key": "debts", "label": "Долги"},
            {"key": "average_price", "label": "Средняя цена"},
            {"key": "remaining_goods", "label": "Остаток товара"},
            {"key": "added_liters", "label": "Поступление"},
        ],
        "rows": [
            {
                "date_label": row["date_label"],
                "liters": _format_number(row["liters"]),
                "amount": _format_number(row["amount"]),
                "payments": _format_number(row["payments"]),
                "debts": _format_number(row["debts"]),
                "average_price": _format_number(row["average_price"]),
                "remaining_goods": _format_number(row["remaining_goods"]),
                "added_liters": _format_number(row["added_liters"]),
            }
            for row in rows_data
        ],
        "totals_row": {
            "date_label": "ИТОГО",
            "liters": _format_number(totals["liters"]),
            "amount": _format_number(totals["amount"]),
            "payments": _format_number(totals["payments"]),
            "debts": _format_number(totals["debts"]),
            "average_price": "—",
            "remaining_goods": "—",
            "added_liters": "—",
        },
        "error_message": error_message,
    }


def _build_report_payload(report_type, start_date, end_date):
    if report_type == "sales":
        return _build_sales_report_payload(start_date, end_date)
    if report_type == "payments":
        return _build_payments_report_payload(start_date, end_date)
    if report_type == "cash":
        return _build_cash_report_payload(start_date, end_date)
    if report_type == "turnover":
        return _build_turnover_report_payload(start_date, end_date)
    return None


def _report_excel_file(report_type, month_start, payload):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Отчет"
    columns = payload.get("columns", [])
    rows = payload.get("rows", [])
    totals_row = payload.get("totals_row")
    total_columns = max(1, len(columns))

    title = f"{REPORT_TYPE_LABELS[report_type]} — {_month_label(month_start)}"
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    title_cell = sheet.cell(row=1, column=1, value=title)
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    thin_side = Side(style="thin", color="000000")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    header_row_index = 3
    for idx, column in enumerate(columns, start=1):
        cell = sheet.cell(row=header_row_index, column=idx, value=column["label"])
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    row_index = header_row_index + 1
    for row in rows:
        for col_idx, column in enumerate(columns, start=1):
            cell = sheet.cell(row=row_index, column=col_idx, value=row.get(column["key"], ""))
            cell.border = border
        row_index += 1

    if totals_row:
        for col_idx, column in enumerate(columns, start=1):
            cell = sheet.cell(row=row_index, column=col_idx, value=totals_row.get(column["key"], ""))
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            cell.border = border

    for col_idx in range(1, total_columns + 1):
        max_length = 0
        for row_cells in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=col_idx, max_col=col_idx):
            value = row_cells[0].value
            if value is None:
                continue
            max_length = max(max_length, len(str(value)))
        sheet.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 60)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"Отчет_{REPORT_TYPE_LABELS[report_type]}_{_month_label_filename(month_start)}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/sales-journal")
@login_required
def sales_journal():
    start_date, end_date = _journal_month_bounds()
    period_label = _journal_period_label(start_date, end_date)
    start_dt, end_dt = _month_datetime_bounds(start_date, end_date)
    q = request.args.get("q", "").strip()
    query = (
        Sale.query
        .options(joinedload(Sale.car).joinedload(Car.client), joinedload(Sale.created_by_user))
        .join(Sale.car)
        .join(Car.client)
        .filter(
            Sale.created_at >= start_dt,
            Sale.created_at < end_dt,
            Car.is_deleted.is_(False),
            Client.is_deleted.is_(False),
        )
    )
    if q:
        query = query.filter(
            or_(
                Client.fio.ilike(f"%{q}%"),
                Car.number.ilike(f"%{q}%"),
            )
        )
    all_sales = query.order_by(Sale.created_at.desc()).all()
    return render_template(
        "sales_journal.html",
        sales=all_sales,
        q=q,
        period_label=period_label,
    )


@app.route("/edit-sale/<int:sale_id>", methods=["GET", "POST"])
@admin_required
def edit_sale(sale_id):
    sale = Sale.query.get_or_404(sale_id)
    errors = {}
    if request.method == "POST":
        try:
            old_liters = _to_decimal_2(sale.liters)
            old_date = _local_date(sale.created_at) or datetime.now(TZ).date()

            liters = _to_decimal_2(request.form["liters"])
            price = _to_decimal_2(request.form["price_per_liter"])
            payment_amount = _to_decimal_2(request.form.get("payment_amount", 0))
            payment_method = request.form["payment_method"]
            if payment_method not in PAYMENT_METHODS:
                raise ValueError("Неверный способ оплаты")

            total = _to_decimal_2(liters * price)
            if payment_method != "долг" and payment_amount > total:
                raise ValueError("Сумма оплаты не может превышать итог продажи")

            sale.liters = liters
            sale.price_per_liter = price
            sale.total = total
            sale.payment_amount = payment_amount
            sale.payment_method = payment_method
            sale.note = request.form.get("note", "") or None
            created_at_str = request.form.get("created_at")
            parsed_created_at = _parse_local_datetime(created_at_str)
            if parsed_created_at:
                sale.created_at = parsed_created_at

            new_date = _local_date(sale.created_at) or datetime.now(TZ).date()
            _adjust_daily_stock(old_liters, old_date)
            _adjust_daily_stock(-liters, new_date)
            _sync_sale_payments_after_edit(sale)

            db.session.commit()
            return redirect(url_for("sales_journal"))
        except (ValueError, TypeError, InvalidOperation) as e:
            errors["main"] = "Ошибка сохранения: " + str(e)
    return render_template("edit_sale.html", sale=sale, errors=errors)


@app.route("/edit-payment/<int:payment_id>", methods=["GET", "POST"])
@admin_required
def edit_payment(payment_id):
    payment = Payment.query.get_or_404(payment_id)
    errors = {}
    if request.method == "POST":
        try:
            payment.amount = float(request.form["amount"])
            payment.payment_method = request.form["payment_method"]
            created_at_str = request.form.get("created_at")
            parsed_created_at = _parse_local_datetime(created_at_str)
            if parsed_created_at:
                payment.created_at = parsed_created_at
            if payment.sale:
                _recalculate_sale_payment_amount(payment.sale)
            db.session.commit()
            return redirect(url_for("payments"))
        except Exception as e:
            errors["main"] = "Ошибка сохранения: " + str(e)
    return render_template("edit_pay.html", payment=payment, errors=errors)


@app.route("/debts-journal")
@login_required
def debts_journal():
    q = request.args.get("q", "").strip()
    client_id = request.args.get("client_id", "").strip()
    query = (
        Sale.query
        .options(joinedload(Sale.car).joinedload(Car.client))
        .join(Sale.car)
        .join(Car.client)
        .filter(
            (Sale.total - db.func.coalesce(Sale.payment_amount, 0.0)) > 0,
            Car.is_deleted.is_(False),
            Client.is_deleted.is_(False),
        )
    )
    if q:
        query = query.filter(
            or_(
                Client.fio.ilike(f"%{q}%"),
                Car.number.ilike(f"%{q}%"),
            )
        )
    if client_id:
        try:
            query = query.filter(Client.id == int(client_id))
        except ValueError:
            pass
    all_debts = query.order_by(Sale.created_at.desc()).all()
    return render_template("debts_journal.html", sales=all_debts, q=q, client_id=client_id)


@app.route("/debts-by-client")
@login_required
def debts_by_client():
    if session.get("role") not in ["admin", "operator"]:
        abort(403)

    q = request.args.get("q", "").strip()
    debt_sales = (
        Sale.query
        .options(joinedload(Sale.car).joinedload(Car.client))
        .join(Sale.car)
        .join(Car.client)
        .filter(
            (Sale.total - db.func.coalesce(Sale.payment_amount, 0.0)) > 0,
            Car.is_deleted.is_(False),
            Client.is_deleted.is_(False),
        )
        .all()
    )
    
    clients_map = {}
    for sale in debt_sales:
        client = sale.car.client
        if client.id not in clients_map:
            clients_map[client.id] = {
                "client": client,
                "count": 0,
                "remaining": Decimal("0.00"),
            }
        clients_map[client.id]["count"] += 1
        clients_map[client.id]["remaining"] += _sale_remaining_debt(sale)

    rows = list(clients_map.values())

    if q:
        rows = [r for r in rows if q.lower() in r["client"].fio.lower()]

    rows.sort(key=lambda x: x["remaining"], reverse=True)

    totals = {
        "count": sum(r["count"] for r in rows),
        "remaining": sum(r["remaining"] for r in rows),
    }

    return render_template("debts_by_client.html", rows=rows, totals=totals, q=q)


@app.route("/pay-debt/<int:sale_id>", methods=["POST"])
@login_required
def pay_debt(sale_id):
    sale = Sale.query.options(joinedload(Sale.car).joinedload(Car.client)).get_or_404(sale_id)
    if not sale.car:
        flash("Машина не найдена. Невозможно создать платеж.", "danger")
        return redirect(url_for("debts_journal"))

    client = sale.car.client
    if not client or client.is_deleted:
        flash("Клиент удален. Невозможно создать платеж.", "danger")
        return redirect(url_for("debts_journal"))

    amount_str = request.form.get("amount", "").strip()
    try:
        amount = Decimal(amount_str)
        if amount <= 0:
            raise ValueError
        amount = _to_decimal_2(amount)
    except (ValueError, TypeError, InvalidOperation):
        return redirect(url_for("debts_journal"))

    payment_method = request.form.get("payment_method", "").strip()
    if payment_method not in ["наличка", "безнал", "доллар"]:
        return redirect(url_for("debts_journal"))

    remaining = _sale_remaining_debt(sale)
    if remaining <= 0:
        flash("Долг по этой продаже уже погашен.", "info")
        client_id = request.form.get("client_id")
        if client_id:
            return redirect(url_for("debts_journal", client_id=client_id))
        return redirect(url_for("debts_journal"))

    if amount > remaining:
        amount = remaining

    payment = Payment(
        client_id=sale.car.client_id,
        sale_id=sale.id,
        amount=amount,
        payment_type="долг",
        payment_method=payment_method,
        paid_by=session.get("user_id"),
    )
    db.session.add(payment)
    db.session.commit()

    # ГЛАВНОЕ ДОБАВЛЕНИЕ:
    client_id = request.form.get("client_id")
    if client_id:
        return redirect(url_for("debts_journal", client_id=client_id))
    else:
        return redirect(url_for("debts_journal"))


@app.route("/payments")
@login_required
def payments():
    start_date, end_date = _journal_month_bounds()
    period_label = _journal_period_label(start_date, end_date)
    start_dt, end_dt = _month_datetime_bounds(start_date, end_date)
    q = request.args.get("q", "").strip()
    query = (
        Payment.query
        .options(
            joinedload(Payment.paid_by_user),
            joinedload(Payment.sale).joinedload(Sale.car),
        )
        .join(Payment.client)
        .filter(Payment.created_at >= start_dt, Payment.created_at < end_dt)
        .order_by(Payment.created_at.desc())
    )
    if q:
        query = query.filter(Client.fio.ilike(f"%{q}%"))
    all_payments = query.all()
    return render_template(
        "payments.html",
        payments=all_payments,
        q=q,
        period_label=period_label,
    )


@app.route("/cash")
@login_required
def cash():
    start_date, end_date = _journal_month_bounds()
    period_label = _journal_period_label(start_date, end_date)
    rows = _build_cash_rows(start_date, end_date)
    return render_template("cash.html", rows=rows, period_label=period_label)


@app.route("/reports")
@login_required
def reports():
    today = datetime.now(TZ).date()
    current_start = today.replace(day=1)
    previous_start = (current_start - timedelta(days=1)).replace(day=1)
    month_options = [
        {
            "value": current_start.strftime("%Y-%m"),
            "label": _month_label(current_start),
        },
        {
            "value": previous_start.strftime("%Y-%m"),
            "label": _month_label(previous_start),
        },
    ]
    return render_template(
        "reports.html",
        default_month=current_start.strftime("%Y-%m"),
        month_options=month_options,
        report_type_labels=REPORT_TYPE_LABELS,
        can_view_turnover=session.get("role") == "admin",
    )


@app.route("/api/reports")
@login_required
def api_reports():
    report_type = request.args.get("report_type", "").strip()
    month_value = request.args.get("month", "").strip()
    start_date, end_date, normalized_month = _month_bounds(month_value)
    if report_type not in REPORT_TYPE_LABELS:
        return jsonify({"error": "Неверный тип отчета."}), 400
    if report_type == "turnover" and session.get("role") != "admin":
        return jsonify({"error": "Доступ запрещен."}), 403

    payload = _build_report_payload(report_type, start_date, end_date)
    response = {
        "report_type": report_type,
        "report_label": REPORT_TYPE_LABELS[report_type],
        "month": normalized_month,
        "month_label": _month_label(start_date),
        "columns": payload.get("columns", []),
        "rows": payload.get("rows", []),
    }
    if payload.get("totals_row"):
        response["totals_row"] = payload["totals_row"]
    if payload.get("error_message"):
        response["error"] = payload["error_message"]
    return jsonify(response)


@app.route("/api/export-report")
@login_required
def export_report():
    report_type = request.args.get("report_type", "").strip()
    month_value = request.args.get("month", "").strip()
    start_date, end_date, _ = _month_bounds(month_value)
    if report_type not in REPORT_TYPE_LABELS:
        return jsonify({"error": "Неверный тип отчета."}), 400
    if report_type == "turnover" and session.get("role") != "admin":
        return jsonify({"error": "Доступ запрещен."}), 403
    payload = _build_report_payload(report_type, start_date, end_date)
    return _report_excel_file(report_type, start_date, payload)


@app.route("/turnover")
@admin_required
def turnover():
    start_date, end_date = _journal_month_bounds()
    period_label = _journal_period_label(start_date, end_date)
    rows_data, totals, error_message = _build_turnover_rows(start_date, end_date)

    return render_template(
        "turnover.html",
        rows=rows_data,
        totals=totals,
        error_message=error_message,
        start_date=start_date.isoformat() if start_date else "",
        end_date=end_date.isoformat() if end_date else "",
        period_label=period_label,
    )


if __name__ == "__main__":
    app.run(debug=os.getenv("FLASK_DEBUG", "False") == "True")
